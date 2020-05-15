import os
import re
from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def process_results_to_xlsx(
        results_dict: dict,
        output_type: str = "file",
        save_file_name: str = "output.xlsx",
        save_file_path: str = "./"
) -> bytes:
    frames = {}
    df = pd.DataFrame.from_dict(results_dict, orient="index")
    replace_column_names = {
        "brand": "Бренд",
        "desr": "Наименование",
        "actionPrice": "Цена по акции",
        "basePrice": "Базовая цена",
        "economy": "Разница",
        "category": "Категория"
    }
    df.rename(columns=replace_column_names, inplace=True)
    all_not_actions = df.drop(["actions"], axis=1).keys().tolist()
    df = df.join(pd.get_dummies(df["actions"].explode()).groupby(level=0).sum())
    frames["Все акции"] = df.drop(["actions", "link"], axis=1).replace(to_replace=0, value='')
    device_links = ("https://mvideo.ru/products/" + df["link"]).to_dict()
    all_actions = df["actions"].explode().dropna().unique().tolist()
    for action in all_actions:
        df_name = str(action)
        df_cat_df = df[all_not_actions].drop(["link"], axis=1)[df[action] == 1]
        df_cat_df = df_cat_df.loc[:, (df_cat_df != 0).any(axis=0)].replace(to_replace=0, value='')
        frames[df_name] = df_cat_df

    wb = Workbook()
    for sheet_name, sheet_df in frames.items():
        sheet_name_escaped = re.sub(r'[^\d\w\s]', '_', str(sheet_name))[:31]
        if len(wb.worksheets) == 1 and wb.worksheets[0].max_row == 1:
            ws = wb.active
            ws.title = sheet_name_escaped
        else:
            ws = wb.create_sheet(sheet_name_escaped, len(wb.worksheets))

        # Data
        for i, r in enumerate(dataframe_to_rows(sheet_df, index=True, header=True)):
            # With index=True, header=True second row is empty, this is a workaround
            if i == 1:
                continue

            # Convert indexes (articles) from str to int
            try:
                if r[0]:
                    r[0] = int(r[0])
            except:
                pass

            # Writing data
            ws.append(r)

            # Writing urls to device descriptions
            row_url = device_links.get(str(r[0]), None)
            if row_url:
                ws.cell(i, 3).hyperlink = row_url
                ws.cell(i, 3).style = "Hyperlink"

        # Headers
        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        # Columns' widths
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
        for col, value in dims.items():
            ws.column_dimensions[col].width = round(value + 10)
        ws.column_dimensions["A"].width = 25

        # Autofilter
        ws.auto_filter.ref = ws.dimensions

        # Freeze panes
        ws.freeze_panes = "D2"

    # Contents sheet
    contents_ws = wb.create_sheet("Акции", 0)
    contents_ws.sheet_properties.tabColor = "1072BA"

    data = [(sheet_name, re.sub(r'[^\d\w\s]', '_', str(sheet_name))[:31]) for sheet_name in frames]
    # print(data)
    contents_ws.cell(row=1, column=1, value="Акция").style = "Pandas"
    contents_ws.cell(row=1, column=2, value="Всего приборов").style = "Pandas"
    contents_ws.cell(row=1, column=3, value="Приборов БСХ").style = "Pandas"
    contents_ws.column_dimensions["A"].width = 50
    contents_ws.column_dimensions["B"].width = 20
    contents_ws.column_dimensions["C"].width = 20

    for y in range(0, len(data)):
        current_cell = contents_ws.cell(row=y + 2, column=1, value=data[y][0])
        current_cell.style = 'Hyperlink'
        cur_df = frames[data[y][0]]
        contents_ws.cell(row=y + 2, column=2, value=cur_df.shape[0])
        contents_ws.cell(row=y + 2, column=3,
                         value=cur_df[(cur_df["Бренд"] == "Bosch") | (cur_df["Бренд"] == "Siemens")].shape[0])
        hyperlink = f"#'{data[y][1]}'!A1"
        current_cell.hyperlink = hyperlink
        referred_cell = wb[data[y][1]].cell(1, 1)
        referred_cell.style = 'Hyperlink'
        referred_cell.value = "К списку акций"
        referred_cell.hyperlink = f"#'Акции'!{current_cell.coordinate}"

    if output_type == "stream":
        with NamedTemporaryFile(delete=False) as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()
            try:
                tmp.close()
                os.unlink(tmp.name)
            except:
                pass
        return stream
    elif output_type == "file":
        save_path = os.path.abspath(os.path.join(save_file_path, save_file_name))
        wb.save(save_path)
        return save_path
    else:
        raise ValueError(f'"output_type" should be "stream" or "file", {output_type} was provided.')


if __name__ == "__main__":
    import json
    with open("G:/serv/dns/mvideo/output.json", "r", encoding="utf8") as f:
        dict_to_save = json.loads(f.read())
    filepath = process_results_to_xlsx(dict_to_save, save_file_name="test.xlsx")
    os.startfile(filepath)